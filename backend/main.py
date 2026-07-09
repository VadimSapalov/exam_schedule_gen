from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, status
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
import openpyxl
from io import BytesIO
from typing import List, Optional

from openpyxl.utils import get_column_letter

from pydantic import BaseModel
from fastapi.responses import StreamingResponse

from scheduler import BacktrackingScheduler
import models
import schemas
from database import engine, get_db
from datetime import date

# Автоматичне створення таблиць
models.Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="Exam Scheduler API Core",
    description="Ядро інформаційної системи автоматизації розкладу сесій",
    version="1.2"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/", tags=["Root"])
def root():
    return {"status": "online", "message": "Exam Scheduler Backend працює стабільно."}


# --- МАРШРУТИ ДЛЯ РОБОТИ З ГРУПАМИ ---

@app.get("/api/v1/references/groups", response_model=List[schemas.GroupNested], tags=["Groups"])
def get_all_groups(db: Session = Depends(get_db)):
    return db.query(models.Group).all()


# --- КОМПЛЕКСНИЙ ІМПОРТ EXCEL (ПРОЦЕС Р1) ---

@app.post("/api/v1/import/bulk/{session_id}", tags=["Import"])
async def import_bulk_system_data(session_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    
    #Комплексний імпорт даних сесії з ОДНОГО Excel-файлу (має містити 5 вкладок: groups, teachers, subjects, rooms, exams_load)
    filename = str(file.filename)
    if not filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="Некоректний формат файлу. Очікується .xlsx або .xls"
        )

    session_entity = db.query(models.SessionEntity).filter(models.SessionEntity.id == session_id).first()
    if not session_entity:
        raise HTTPException(status_code=404, detail=f"Сесію з ID {session_id} не знайдено в системі. Створіть сесію перед імпортом.")

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
        summary = {}
        errors = []

        # --- ОЧИЩЕННЯ ТАБЛИЦЬ ПЕРЕД ІМПОРТОМ ---
        try:
            # 1. Видаляємо залежні дані (чернетки та плани іспитів)
            db.query(models.ScheduleDraft).delete()
            db.query(models.Exam).delete() 
            db.commit()

            # 2. Видаляємо головні довідники
            db.query(models.Group).delete()
            db.query(models.Teacher).delete()
            db.query(models.Subject).delete()
            db.query(models.Room).delete()
            db.commit()
        except Exception as clean_err:
            db.rollback()
            raise HTTPException(
                status_code=500, 
                detail=f"Помилка підготовки БД (очищення старих даних): {str(clean_err)}"
            )

        # 1. ПАРСИНГ ЛИСТА: groups
        if "groups" in workbook.sheetnames:
            sheet = workbook["groups"]
            g_count = 0
            
            # Локальні лічильники для шахового чергування в межах одного імпорту
            bachelor_idx = 0
            master_idx = 0
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or row[0] is None: continue
                
                # Читаємо тільки потрібні 3 колонки, ігноруючи все інше
                code = str(row[0])
                level = str(row[1]).lower().strip()
                student_count = row[2]
                
                if level not in ["bachelor", "master"]:
                    errors.append(f"Лист 'groups', рядок {row_idx}: Невалідний рівень навчання '{level}'. Очікується 'bachelor' або 'master'.")
                    continue
                
                # АВТОМАТИЧНЕ ШАХОВЕ ЧЕРГУВАННЯ
                if level == "bachelor":
                    # Парні бакалаврські групи йдуть на Пн/Ср/Пт, непарні — на Вт/Чт
                    auto_day_pattern = "mon_wed_fri" if bachelor_idx % 2 == 0 else "tue_thu"
                    bachelor_idx += 1
                else:
                    # Магістрів пускаємо дзеркально, щоб мінімізувати перетини на старті
                    auto_day_pattern = "tue_thu" if master_idx % 2 == 0 else "mon_wed_fri"
                    master_idx += 1
                
                db_group = db.query(models.Group).filter(models.Group.code == code).first()
                if not db_group:
                    val_student_count = 0
                    if isinstance(student_count, (int, float)):
                        val_student_count = int(student_count)
                    elif student_count is not None:
                        try:
                            val_student_count = int(float(str(student_count)))
                        except (ValueError, TypeError):
                            val_student_count = 0

                    db_group = models.Group(
                        code=code, 
                        level=level, 
                        student_count=val_student_count, 
                        day_pattern=auto_day_pattern
                    )
                    db.add(db_group)
                    g_count += 1
            summary["groups_imported"] = g_count

        # 2. ПАРСИНГ ЛИСТА: teachers
        if "teachers" in workbook.sheetnames:
            sheet = workbook["teachers"]
            t_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or row[0] is None: continue
                
                full_name = str(row[0])
                # Безпечно беремо додаткові змінні, якщо вони є в Excel, або ставимо дефолт
                position = str(row[1]) if len(row) > 1 and row[1] is not None else "Викладач"
                is_active = row[2] if len(row) > 2 and row[2] is not None else True
                
                db_teacher = db.query(models.Teacher).filter(models.Teacher.full_name == full_name).first()
                if not db_teacher:
                    db_teacher = models.Teacher(full_name=full_name, position=position, is_active=bool(is_active))
                    db.add(db_teacher)
                    t_count += 1
            summary["teachers_imported"] = t_count

        # 3. ПАРСИНГ ЛИСТА: subjects
        if "subjects" in workbook.sheetnames:
            sheet = workbook["subjects"]
            s_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or row[0] is None: continue
                
                name = str(row[0])
                # Безпечно перевіряємо наявність колонок
                requires_computer = row[1] if len(row) > 1 and row[1] is not None else False
                level = str(row[2]).lower().strip() if len(row) > 2 and row[2] is not None else "bachelor"
                
                db_sub = db.query(models.Subject).filter(models.Subject.name == name, models.Subject.level == level).first()
                if not db_sub:
                    db_sub = models.Subject(name=name, requires_computer=bool(requires_computer), level=level)
                    db.add(db_sub)
                    s_count += 1
            summary["subjects_imported"] = s_count

        # 4. ПАРСИНГ ЛИСТА: rooms
        if "rooms" in workbook.sheetnames:
            sheet = workbook["rooms"]
            r_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or row[0] is None: continue
                name, capacity, is_computer_lab, building = str(row[0]), row[1], row[2], str(row[3])
                
                db_room = db.query(models.Room).filter(models.Room.name == name).first()
                if not db_room:
                    val_capacity = 0
                    if isinstance(capacity, (int, float)):
                        val_capacity = int(capacity)
                    elif capacity is not None:
                        try:
                            val_capacity = int(float(str(capacity)))
                        except (ValueError, TypeError):
                            val_capacity = 0

                    db_room = models.Room(name=name, capacity=val_capacity, is_computer_lab=bool(is_computer_lab), building=building)
                    db.add(db_room)
                    r_count += 1
            summary["rooms_imported"] = r_count

        db.commit()

        # 5. ПАРСИНГ ЛИСТА: exams_load (Формування плану іспитів сесії)
        if "exams_load" in workbook.sheetnames:
            sheet = workbook["exams_load"]
            e_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or row[0] is None: continue
                group_code, subject_name, teacher_name = str(row[0]), str(row[1]), str(row[2])
                
                g = db.query(models.Group).filter(models.Group.code == group_code).first()
                t = db.query(models.Teacher).filter(models.Teacher.full_name == teacher_name).first()
                s = db.query(models.Subject).filter(models.Subject.name == subject_name).first()
                
                if not g or not t or not s:
                    errors.append(f"Лист 'exams_load', рядок {row_idx}: Не знайдено зв'язків у базі для групи '{group_code}', предмета '{subject_name}' або викладача '{teacher_name}'.")
                    continue
                
                # Перевіряємо чи такий іспит уже додано до цієї сесії
                existing_exam = db.query(models.Exam).filter(
                    models.Exam.session_id == session_id,
                    models.Exam.group_id == g.id,
                    models.Exam.subject_id == s.id
                ).first()
                
                if not existing_exam:
                    new_exam = models.Exam(group_id=g.id, subject_id=s.id, teacher_id=t.id, session_id=session_id)
                    db.add(new_exam)
                    e_count += 1
            summary["exams_loaded"] = e_count

        db.commit()
        return {"status": "success", "imported_data": summary, "errors": errors}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка комплексного імпорту: {str(e)}")


# --- МОДУЛЬ АВТОМАТИЧНОЇ ГЕНЕРАЦІЇ ТА ЧЕРНЕТОК РОЗКЛАДУ ---

class GenerateSchedulePayload(BaseModel):
    mode: str

@app.post("/api/v1/schedule/generate/{session_id}", tags=["Schedule"])
def generate_schedule(session_id: int, payload: GenerateSchedulePayload, db: Session = Depends(get_db)):
    session_entity = db.query(models.SessionEntity).filter(models.SessionEntity.id == session_id).first()
    if not session_entity:
        raise HTTPException(status_code=404, detail="Зазначену сесію не знайдено")
        
    # Автоматично оновлюємо режим у сесії, якщо він прийшов інший з фронту
    if payload.mode in ["teacher_density", "room_density"]:
        setattr(session_entity, "optimization_mode", payload.mode)
        db.commit()

    db.query(models.ScheduleDraft).filter(models.ScheduleDraft.session_id == session_id).delete()
    db.commit()

    # Передаємо обраний режим (payload.mode) в архітектуру планувальника
    scheduler = BacktrackingScheduler(db=db, session_id=session_id)
    result = scheduler.generate()
    
    if result is None:
        raise HTTPException(
            status_code=422, 
            detail="Неможливо побудувати розклад без колізій жорстких обмежень. Недостатньо вільних днів або викладачів."
        )
        
    try:
        for item in result:
            slot_date = item["slot"]["date"]
            slot_letter = item["slot"]["slot"]
            
            db_slot = db.query(models.ScheduleSlot).filter(
                models.ScheduleSlot.session_id == session_id,
                models.ScheduleSlot.exam_date == slot_date,
                models.ScheduleSlot.slot == slot_letter
            ).first()
            
            if not db_slot:
                db_slot = models.ScheduleSlot(
                    session_id=session_id,
                    exam_date=slot_date,
                    slot=slot_letter,
                    week_number=item["slot"]["week"]
                )
                db.add(db_slot)
                db.commit()
                db.refresh(db_slot)

            draft_entry = models.ScheduleDraft(
                exam_id=item["exam"].id,
                slot_id=db_slot.id,
                room_id=item["room"].id,
                session_id=session_id,
                status="draft",
                is_manual_edit=False,
                conflict_flag=False
            )
            db.add(draft_entry)
            
        db.commit()
        return {"status": "success", "message": f"Розклад успішно згенеровано. Розподілено іспитів: {len(result)}"}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка фіксації розкладу в БД: {str(e)}")


@app.get("/api/v1/schedule/drafts/{session_id}", response_model=List[schemas.ScheduleDraftResponse], tags=["Schedule"])
def get_schedule_drafts(session_id: int, db: Session = Depends(get_db)):
    #Отримання списку чернеток розкладу для відображення на фронтенді
    drafts = db.query(models.ScheduleDraft).filter(models.ScheduleDraft.session_id == session_id).all()
    return drafts

# --- МАРШРУТИ ДЛЯ РОБОТИ З СЕСІЯМИ ---

@app.get("/api/v1/sessions/", tags=["Sessions"])
def get_all_sessions(db: Session = Depends(get_db)):
    #Отримання списку всіх екзаменаційних сесій для перемикача на фронтенді
    sessions = db.query(models.SessionEntity).all()
    # Якщо база порожня, створимо початкові сесії автоматично для зручності тестування
    if not sessions:
        session1 = models.SessionEntity(name="Літня екзаменаційна сесія 2026", optimization_mode="teacher_density")
        session2 = models.SessionEntity(name="Зимова екзаменаційна сесія 2026", optimization_mode="room_density")
        db.add(session1)
        db.add(session2)
        db.commit()
        sessions = [session1, session2]
    return sessions

from pydantic import BaseModel
class SessionUpdateMode(BaseModel):
    optimization_mode: str

@app.patch("/api/v1/sessions/{session_id}", tags=["Sessions"])
def update_session_optimization_mode(session_id: int, payload: SessionUpdateMode, db: Session = Depends(get_db)):
    #Оновлення критерію оптимізації сесії при зміні селектора на фронтенді
    session_entity = db.query(models.SessionEntity).filter(models.SessionEntity.id == session_id).first()
    if not session_entity:
        raise HTTPException(status_code=404, detail="Сесію не знайдено")
    
    if payload.optimization_mode not in ["teacher_density", "room_density"]:
        raise HTTPException(status_code=400, detail="Невалідний режим оптимізації. Доступні: teacher_density, room_density")
        
    setattr(session_entity, "optimization_mode", payload.optimization_mode)
    db.commit()
    db.refresh(session_entity)
    return session_entity

# --- 1. СХЕМА ДЛЯ РЕДАГУВАННЯ ЧЕРНЕТКИ ---
class EditDraftPayload(BaseModel):
    slot_date: date
    slot_letter: str
    room_id: Optional[int] = None  # Тепер це необов'язково

# --- ЕНДПОІНТ РЕДАГУВАННЯ ---
@app.patch("/api/v1/schedule/drafts/{draft_id}", tags=["Schedule"])
def edit_schedule_draft(draft_id: int, payload: EditDraftPayload, db: Session = Depends(get_db)):

    #Редагування запису чернетки розкладу з перевіркою зайнятості аудиторії.
    draft = db.query(models.ScheduleDraft).filter(models.ScheduleDraft.id == draft_id).first()
    if not draft:
        raise HTTPException(status_code=404, detail="Запис чернетки не знайдено")

    # Знаходимо цільовий слот
    target_slot = db.query(models.ScheduleSlot).filter(
        models.ScheduleSlot.exam_date == payload.slot_date,
        models.ScheduleSlot.slot == payload.slot_letter
    ).first()

    if not target_slot:
        raise HTTPException(
            status_code=400, 
            detail=f"Слот {payload.slot_letter} на {payload.slot_date} не знайдено в базі"
        )

    # Визначаємо аудиторію (нову або залишаємо поточну)
    target_room_id = payload.room_id if payload.room_id is not None else draft.room_id

    # Сувора валідація унікальності uq_room_per_slot перед збереженням
    overlapping_draft = db.query(models.ScheduleDraft).filter(
        models.ScheduleDraft.slot_id == target_slot.id,
        models.ScheduleDraft.room_id == target_room_id,
        models.ScheduleDraft.id != draft_id
    ).first()

    if overlapping_draft:
        raise HTTPException(
            status_code=400,
            detail="Ця аудиторія в зазначеному слоті вже зайнята іншим іспитом!"
        )

    # Оновлюємо координати чернетки
    draft.slot_id = target_slot.id  # type: ignore
    if payload.room_id is not None:
        draft.room_id = payload.room_id  # type: ignore
        
    draft.is_manual_edit = True  # type: ignore
    draft.conflict_flag = False  # type: ignore

    db.commit()
    return {"status": "success", "message": "Дані чернетки успішно оновлено"}
# --- 3. ЕНДПОІНТ ЕКСПОРТУ В EXCEL ---
@app.get("/api/v1/schedule/export/{session_id}", tags=["Schedule"])
def export_schedule_to_excel(session_id: int, db: Session = Depends(get_db)):
    
    # Гнучкий експорт сітки розкладу в Excel (дозволяє вивантаження навіть з конфліктами).
    drafts = db.query(models.ScheduleDraft).filter(models.ScheduleDraft.session_id == session_id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None 
    
    ws.title = "Розклад сесії"

    headers = ["Тиждень", "Дата", "Слот", "Група", "Предмет", "Викладач", "Аудиторія", "Тип Аудиторії"]
    ws.append(headers)

    for d in drafts:
        week = d.slot.week_number if d.slot else ""
        exam_date = d.slot.exam_date.strftime("%Y-%m-%d") if d.slot and d.slot.exam_date else ""
        slot_letter = d.slot.slot if d.slot else ""
        group = d.exam.group.code if d.exam and d.exam.group else ""
        subject = d.exam.subject.name if d.exam and d.exam.subject else ""
        teacher = d.exam.teacher.full_name if d.exam and d.exam.teacher else ""
        room = d.room.name if d.room else ""
        room_type = "Лабораторія" if d.room and d.room.is_computer_lab else "Лекційна"
        
        is_manual = bool(d.is_manual_edit)
        is_conflict = bool(d.conflict_flag)
        #edit_status не експортується в файл
        edit_status = "Вручну" if is_manual else "Автоматично"

        if is_conflict:
            edit_status += " (Є колізія!)"
        
        ws.append([week, exam_date, slot_letter, group, subject, teacher, room, room_type])

    # Авто-ширина колонок
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        # Використовуємо функцію, яку імпортували напряму
        assert col[0].column is not None
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=schedule_session_{session_id}.xlsx"}
    )